import eurostat
import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from utils.calculation.calculate_indicator import indicator_avg, indicator_inverse


def collect_data(code: str, countries: list, size_emp: str, ind_type: str, indic_is: list, unit: str,
                 startPeriod: str, endPeriod: str, ind_id: str):
    """
    Function to collect data from eurostat using eurostat's API

    :param code: str, the eurostat code of the table
    :param countries: list, the ISO-2 codes of the countries
    :param size_emp: str, the eurostat code for the size of enterprise
    :param ind_type: str, the eurostat code for the type of indicator
    :param indic_is: list, the eurostat codes of the variables of the indicator
    :param unit: the eurostat code for the unit of measure of the variables
    :param startPeriod: str, the starting year of interest
    :param endPeriod: str, the ending year of interest
    :param ind_id: str, internal indicator id
    :return: pandas dataframe with columns ['freq',	'size_emp',	'nace_r2', 'indic_is', 'unit', 'geo\TIME_PERIOD',
            '2022'] where the 2022 is the values for year 2022. All variables are under indic_is column

    """
    my_filter_pars = {'size_emp': size_emp,
                      'ind_type': ind_type,
                      'indic_is': indic_is,
                      'unit': unit,
                      'geo': countries,
                      'startPeriod': startPeriod,
                      'endPeriod': endPeriod}

    logging.info(f"Data collection for indicator {ind_id} has started..")

    data = None
    var_names = None

    try:
        data = eurostat.get_data_df(code, filter_pars=my_filter_pars)
        var_names = eurostat.get_dic(code=code, par='indic_is')
    except Exception as e:
        logging.error(f"Data for indicator {ind_id} were not collected due to {e}")

    return data, dict(var_names)


def write_data(data, original_data, path: str, ind_id: str = None, sheet_name: str = None, computation: str = None,
               indicator_name: str = None):
    """Function to write the data for one indicator in one excel, with different tab for each variable.
    :param original_data: data as collected from eurostat API
    :param data: pandas dataframe with the form resulted from collect_data()
    :param path:  str, the path to write the Excel file to
    :param ind_id: str, the internal indicator id
    :param sheet_name: str, the name of the Sheet
    :param computation: if indicator's value needs to be computed
    :param indicator_name: indicator's name, needed when indicator has been computed
    """
    if not os.path.exists(path):

        with pd.ExcelWriter(path) as writer:

            for i, var in enumerate(np.append(original_data['indic_is'].unique(), indicator_name if computation == "average" else None)):
                if not sheet_name:
                    ind_id = str(ind_id).replace('.', '')
                    sheet_name = f'{ind_id}_{i + 1}_{var}'

                if var:
                    if computation == "average":
                        data[["freq", "size_emp", "nace_r2", "unit", "geo\TIME_PERIOD", var]].to_excel(writer, sheet_name=sheet_name, index=False)
                    else:
                        data.loc[data['indic_is'] == var].to_excel(writer, sheet_name=sheet_name, index=False)

                sheet_name = None

        logging.info(f"Data have been written to {path}.")

    else:
        logging.error(f'File {path} already exists, please rename or delete.')
        pass


def write_data_ki_schema(data, path, ind_id=None, sheet_name=None):
    """Function to open existing Excel and write each variable in a new tab"""
    book = load_workbook(path)
    ws = book.create_sheet(sheet_name)

    for r in dataframe_to_rows(data, index=True, header=True):
        ws.append(r)

    book.save(path)

    logging.info(f"Data have been written to {path}.")


def _read_data(path):
    """Function to read data from Excel file specified in the path. This will load the file as a dictionary of
    dataframes. """
    data = pd.read_excel(path, sheet_name=None)

    return data


def consolidate_data(data_dir: str):
    """Function to read data from specified path, consolidate them into one dataframe
    :param data_dir: str, the directory to read all excel files from
    :return pandas dataframe with the form ['freq',	'size_emp',	'nace_r2', 'unit', 'geo\TIME_PERIOD', 'variable_name_1',
            variable_name_2', ... ] for all files in data_dir
    """

    files = os.listdir(data_dir)
    all_data_vars = None
    all_data = None

    for j, file in enumerate(files):
        data = _read_data(data_dir + file)

        for i, key in enumerate(data.keys()):
            temp = data.get(key)
            temp.rename(columns={temp.columns[-1]: key}, inplace=True)

            if i == 0:
                all_data_vars = temp
            else:
                all_data_vars = all_data_vars.merge(data.get(key)[['geo\TIME_PERIOD', key]], on='geo\TIME_PERIOD',
                                                    how='inner')

        if j == 0:
            all_data = all_data_vars
        else:
            cols = ['geo\TIME_PERIOD'] + list(data.keys())

            all_data = all_data.merge(all_data_vars[cols], on='geo\TIME_PERIOD',
                                      how='inner')

    return all_data


def calculate_indicators(data, cols: list, computation_alg: str, name: str, max_value=None):
    """
    Function to calculate the values of some indicators from their variables
    :param data:
    :param cols: the columns to use for the computation (variable columns)
    :param computation_alg: the computation algorithm to use
    :param name: internal indicator's name
    :param max_value: the value to use for the inverse
    :return: pandas dataframe with an extra column as the indicator value
    """
    if computation_alg == 'average':
        temp = data.pivot(index=['freq', 'size_emp', 'nace_r2', 'unit', 'geo\TIME_PERIOD'], columns='indic_is',
                          values='2022').reset_index().copy()
        temp[name] = indicator_avg(temp, cols)
        return temp, data
    elif computation_alg == 'inverse':
        temp = data.rename(columns={data.columns[-1]: cols[0]}).copy()
        temp[name] = indicator_inverse(temp, cols[0], max_value)
        return temp, data

    else:
        logging.info("Computation algorithm not one of average, inverse...")
        pass
